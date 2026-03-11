import os
from openpyxl import Workbook 
import pdfplumber
import re
from datetime import datetime

# 1 importa classe para criar planilhas Excel
# 2 biblioteca que abre e lê texto de PDFs
# 3 biblioteca de expressões regulares (usada para encontrar padrões de texto)
# 4 biblioteca para trabalhar com datas e horas


def main():

    # nome da pasta onde estão os PDFs
    directory = 'pdf_invoice'

    # lista todos os arquivos dentro da pasta
    files = os.listdir(directory)

    # conta quantos arquivos existem na pasta
    files_quantily = len(files)

    # se não houver arquivos, o programa para e mostra erro
    if files_quantily == 0:
        raise Exception("No files found in the directory")

    # cria um novo arquivo Excel
    wb = Workbook()

    # seleciona a planilha ativa
    ws = wb.active

    # define o nome da aba da planilha
    ws.title = 'Invoice Imports'

    # cria os cabeçalhos das colunas na planilha
    ws['A1'] = 'Invoice #'
    ws['B1'] = 'Date'
    ws['C1'] = 'File Name'
    ws['D1'] = 'Status'

    # define que os dados começarão na linha 2 
    last_empty_line = 2

    # percorre por todos os arquivos da pasta
    for file in files:

        # abre o PDF
        with pdfplumber.open(directory + "/" + file) as pdf:

            # pega a primeira página do PDF
            first_page = pdf.pages[0]

            # extrai todo o texto dessa página
            pdf_text = first_page.extract_text()

        # mostra no terminal qual arquivo está sendo processado
        print("ARQUIVO:", file)

        # mostra o texto extraído do PDF
        print(pdf_text)
        print("-" * 50)

        inv_number_re_pattern = r'INVOICE #(\d+)'

        inv_date_re_pattern = r'DATE:\s*(\d{2}/\d{2}/\d{4})'

        # procura o número da invoice dentro do texto
        match_number = re.search(inv_number_re_pattern, pdf_text)

        # procura a data dentro do texto
        match_date = re.search(inv_date_re_pattern, pdf_text)

        # se encontrou o número da invoice
        if match_number:

            # pega apenas o número encontrado
            invoice_number = match_number.group(1)

            # escreve o número na coluna A
            ws[f'A{last_empty_line}'] = invoice_number

        else:
            # se não encontrar, escreve mensagem de erro
            ws[f'A{last_empty_line}'] = "Couldn't find invoice number"

        # se encontrou a data
        if match_date:

            # pega apenas a data encontrada
            invoice_date = match_date.group(1)

            # escreve a data na coluna B
            ws[f'B{last_empty_line}'] = invoice_date

        else:
            # se não encontrar, escreve mensagem de erro
            ws[f'B{last_empty_line}'] = "Couldn't find invoice date"

        # escreve o nome do arquivo PDF na coluna C
        ws[f'C{last_empty_line}'] = file

        # escreve status de processamento
        ws[f'D{last_empty_line}'] = "Completed"

        # passa para a próxima linha da planilha
        last_empty_line += 1

    # pega data e hora atuais
    full_now = str(datetime.now()).replace(":", "-")

    # encontra o ponto antes dos milissegundos
    dot_index = full_now.index(".")

    # corta os segundos da data
    now = full_now[:dot_index]

    # salva o arquivo Excel com data no nome
    wb.save("invoice_report.xlsx")


if __name__ == "__main__":
    main()